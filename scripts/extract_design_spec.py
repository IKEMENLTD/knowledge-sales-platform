"""
sales_platform_design_spec_v2.xlsx の全 sheet を markdown / json に展開する。
gap 分析の入力にする。

usage:
    python scripts/extract_design_spec.py
"""

import json
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
SPEC = Path(
    r"C:\Users\ooxmi\Downloads\営業、CS統合管理システム＿ナレッジさん\sales_platform_design_spec_v2.xlsx"
)
OUT_MD_DIR = REPO / "reviews" / "design_spec_md"
OUT_JSON = REPO / "reviews" / "design_spec.json"


def cell_str(v):
    if v is None:
        return ""
    return str(v).strip()


def main():
    OUT_MD_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(SPEC, read_only=True, data_only=True)

    spec = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [cell_str(c) for c in row]
            if any(cells):
                rows.append(cells)
        spec[sheet_name] = rows

        # md 出力
        max_cols = max((len(r) for r in rows), default=0)
        md = [f"# {sheet_name}", ""]
        for i, row in enumerate(rows):
            padded = row + [""] * (max_cols - len(row))
            md.append("| " + " | ".join(padded) + " |")
            if i == 0:
                md.append("| " + " | ".join(["---"] * max_cols) + " |")
        (OUT_MD_DIR / f"{sheet_name}.md").write_text("\n".join(md), encoding="utf-8")
        print(f"  {sheet_name}: {len(rows)} rows")

    OUT_JSON.write_text(json.dumps(spec, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n  -> {OUT_MD_DIR}")
    print(f"  -> {OUT_JSON}")


if __name__ == "__main__":
    main()
